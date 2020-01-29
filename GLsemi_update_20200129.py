
import openpyxl
import os
import calendar
import time
from tkinter import messagebox


#주의사항

print('###사용시 유의사항###\n\n1. 출하파일에 공백인 행이 없도록 해야함')
print('2. 출하파일 입력시 날짜에 OO 형식으로 년,월 제외하고 일 만 2자릿 수로 작성')
print('3. 모든 파일은 C:\\세계자동화_지엘반도체 폴더 안에 존재해야함\n\n잠시만 기다려주십시오...\n')


#기본 세팅

path = "C:\\세계자동화_GL반도체"
file_names = os.listdir(path)
for name in file_names :
    if str('고객_GL반도체') in name :
        file1 = os.path.join(path, name)
    elif str('출하_GL반도체') in name :
        file2 = os.path.join(path, name)


of1 = openpyxl.load_workbook(file1, data_only=True)
of2 = openpyxl.load_workbook(file2, data_only=True)
of3 = openpyxl.load_workbook(r'C:\세계자동화_GL반도체\세금계산서등록양식일반.xlsx')


sh1 = of1['거래처 목록']
sh2 = of2['출하내역']
sh3 = of3['엑셀업로드양식']



#고정 자료 삽입


na = int(time.strftime('%m', time.localtime(time.time())))
day = calendar.monthrange(2020,na) [1]


total = int(4)
while 1 :
    if sh2.cell(row=total, column=1).value is not None :
        total = total + 1
    else :
        break
if total > 103 :
    messagebox.showwarning("알림", "100건 초과입니다! 100건 이후의 출하내역은 따로 처리 부탁드립니다.")


for i in range(7, total+3):
    sh3.cell(row=i, column=1). value = '01'
    sh3.cell(row=i, column=2). value = '2020'+str(na)+str(day)
    sh3.cell(row=i, column=3). value = '1018666667'
    sh3.cell(row=i, column=5). value = '지엘반도체 주식회사'
    sh3.cell(row=i, column=6). value = '최상영'
    sh3.cell(row=i, column=59). value = '02'




#출하 현황에서 자료 가져오기

for k in range(7, total+3):
    sh3.cell(row=k, column=13).value = sh2.cell(row=(k - 3), column=1).value  # 상호
    sh3.cell(row=k, column=20).value = sh2.cell(row=(k - 3), column=10).value  # 공급가액
    sh3.cell(row=k, column=28).value = sh2.cell(row=(k - 3), column=10).value  # 공급가액1
    sh3.cell(row=k, column=21).value = sh2.cell(row=(k - 3), column=12).value  # 세액
    sh3.cell(row=k, column=29).value = sh2.cell(row=(k - 3), column=12).value  # 세액1
    sh3.cell(row=k, column=23).value = sh2.cell(row=(k - 3), column=4).value  # 일자



#거래처 목록에서 자료 가져오기

for x in range(7,total+3) :
    for y in range(5,833) :
        if sh3.cell(row=x, column=13).value in sh1.cell(row=y, column=4).value :
            sh3.cell(row=x, column=11).value = sh1.cell(row=y, column=2).value
            sh3.cell(row=x, column=14).value = sh1.cell(row=y, column=5).value
            sh3.cell(row=x, column=13).value = sh1.cell(row=y, column=4).value




"""
from openpyxl.styles import Alignment

#정렬하는 법 (안됨..)
sh3.cell(row=30, column=20).alignment = Alignment(horizontal="center", vertical="center")
of2.alignment = Alignment(horizontal="center", vertical="center")
"""

#셀 저장하기

of3.save(r'C:/세계자동화_GL반도체/result.xlsx')

messagebox.showinfo("알림", "result.xlsx 파일이 생성되었습니다.")


